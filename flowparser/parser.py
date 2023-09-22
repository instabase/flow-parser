"""
MIT License

Copyright (c) 2023 Instabase

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
import argparse
import json
import os
import re
from collections import OrderedDict
from datetime import datetime
from operator import itemgetter
from pathlib import Path
import numpy
from jinja2 import Environment, FileSystemLoader
import csv

def parse(logs, waiting_threshold):

    # Determine which time format is used
    try:
        format = "%Y-%m-%d %H:%M:%S,%f"
        datetime.strptime(logs[0]["ts"], format)
    except:
        try:
            format = "%m/%d/%y, %I:%M:%S %p"
            datetime.strptime(logs[0]["ts"], format)
        except:
            #15/09/2023, 18:55:33
            format = '%d/%m/%Y, %H:%M:%S'
            datetime.strptime(logs[0]["ts"], format)

    # Check log timestamp order in case order was modified before saving from UI
    if datetime.strptime(logs[0]["ts"], format) > datetime.strptime(logs[-1]["ts"], format):
        logs.reverse()

    job_details = {}
    job_id = logs[-1]["jobId"]
    job_details.setdefault("jobid", job_id)
    job_details.setdefault(
        "start_time", datetime.strptime(logs[0]["ts"], format))
    job_details.setdefault(
        "end_time", datetime.strptime(logs[-1]["ts"], format))
    job_details.setdefault(
        "elapsed_time",
        (
            datetime.strptime(logs[-1]["ts"], format) -
            datetime.strptime(logs[0]["ts"], format)
        ).total_seconds(),
    )

    stage_to_filename = {}
    stage_to_type = {}
    reduce_stage_to_filename = {}

    # Inital loop of logs to map filenames to steps and filenames to stages/task ids
    for log in logs:
        log_details = log["log"]

        # Kick out any log entry that is over 5000 char. This can cause regex to take a long time if long log elements present
        if len(log_details) > 5000:
            continue

        # Regex step and filename identifiers
        file_name_process_files = re.findall(
            r"^Running process-file on .*\/(.*)\.\w+", log_details
        )
        file_name_output_dir = re.findall(
            r"^Running process-file on (.*)", log_details
        )
        file_name_classifer = re.findall(
            r"^Written file=.*\/(.*).csv successfully.", log_details
        )
        file_name_classifer_2 = re.findall(
            r"^Written file=.*\/(.*)\.\w+\.ibmsg successfully", log_details
        )
        file_name_refiner_split_class = re.findall(
            r"^record=(.*)_.*_.*\.ibmsg column.*", log_details
        )
        file_name_refiner = re.findall(
            r"^record=(.*)\.\w+\.ibmsg column.*", log_details)
        file_name_refiner_22_10 = re.findall(
            r"Refiner Record: (.*)\.\w+ Field: (.*) run time: (.*) sec", log_details
        )

        file_name_reduce_udf = re.findall(
            r"^file_node: (.*)\.ibmsg", log_details)
        file_name_custom_classifer = re.findall(
            r"^filename : (.*)\.", log_details)
        reduce_udf = re.findall(r".*reduce operation", log_details)
        file_name_flow_split = re.findall(
            r"^Inferred process_type of file\s(.*)_pg\d+_pg\d+\.pdf\s.*",
            log_details,
        )
        file_name_flow_v2 = re.findall(
            r"^Inferred process_type of file\s(.*)\.\w+ to be \w+", log_details
        )
        generic_step_identifer = re.findall(
            r"Executing \w+ substep for: (\w+) id: (\w+)", log_details
        )
        if file_name_flow_split:
            stage_to_filename.setdefault(
                log["taskId"], file_name_flow_split[0] + ".pdf"
            )
            stage_to_type.setdefault(log["step"], "process_files")
        if file_name_flow_v2:
            stage_to_filename.setdefault(
                log["taskId"], file_name_flow_v2[0])
            stage_to_type.setdefault(log["step"], "process_files")
        if file_name_process_files:
            stage_to_filename.setdefault(
                log["taskId"], file_name_process_files[0])
            stage_to_type.setdefault(log["step"], "process_files")

        if file_name_output_dir:
            job_details.setdefault(
                "outputDir", "/".join(
                    file_name_output_dir[0].split("/")[:-2])
            )
        if file_name_refiner_22_10:
            stage_to_filename.setdefault(
                log["taskId"], file_name_refiner_22_10[0][0]
            )
            stage_to_type.setdefault(log["step"], "apply_refiner")
        if file_name_classifer:
            stage_to_filename.setdefault(
                log["taskId"], file_name_classifer[0])
            stage_to_type.setdefault(log["step"], "classifer")
        if file_name_classifer_2:
            stage_to_filename.setdefault(
                log["taskId"], file_name_classifer_2[0])
            stage_to_type.setdefault(log["step"], "classifer")
        if file_name_refiner_split_class:
            stage_to_filename.setdefault(
                log["taskId"], file_name_refiner_split_class[0]
            )
            stage_to_type.setdefault(log["step"], "apply_refiner")
        if file_name_refiner:
            stage_to_filename.setdefault(
                log["taskId"], file_name_refiner[0])
            stage_to_type.setdefault(log["step"], "apply_refiner")
        if file_name_reduce_udf:
            stage_to_filename.setdefault(
                log["taskId"], file_name_reduce_udf[0])
            stage_to_type.setdefault(log["step"], "reduce udf")
            reduce_stage_to_filename.setdefault(log["step"], []).append(
                file_name_reduce_udf[0]
            )
        if reduce_udf:
            stage_to_type.setdefault(log["step"], "reduce udf")
        if generic_step_identifer:
            stage_to_type.setdefault(
                generic_step_identifer[0][1], generic_step_identifer[0][0]
            )
        if file_name_custom_classifer:
            stage_to_filename.setdefault(
                log["taskId"], file_name_custom_classifer[0]
            )
            stage_to_type.setdefault(log["step"], "custom classifer")
        if "logPrefix" in log:
            logPrefix = re.findall(
                r"org=(.*) project=(.*) batchguid=(.*)", log["logPrefix"]
            )
            job_details.setdefault("org", logPrefix[0][0])
            job_details.setdefault("project", logPrefix[0][1])
            job_details.setdefault("batch", logPrefix[0][2])

    span_interference_dict = {}
    for log in logs:
        log_details = log["log"]

        # Kick out any log entry that is over 5000 char. This can cause regex to take a long time if long log elements present
        if len(log_details) > 5000:
            continue

        ts = datetime.strptime(log["ts"], format)

        # Run model inference
        if 'Run model inference' in log_details and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })

        # Acquiring lock for ModelID(name=ibllm, version=1.1.19, is_project_model=False, username=system)
        model_acquiring_lock = re.findall(
            r"Acquiring lock for ModelID\(name=(\w+), version=(\d+.\d+.\d+), is_project_model=(\w+), username=system\)", log_details)
        if model_acquiring_lock and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })
        # Got RunModel request for ModelID(name=ibllm, version=1.1.19, is_project_model=False, username=system)
        model_runmodel_request = re.findall(
            r"Got RunModel request for ModelID\(name=(\w+), version=(\d+.\d+.\d+), is_project_model=(\w+), username=system\)", log_details)
        if model_runmodel_request and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            span.setdefault('model_name', model_runmodel_request[0][0])
            span.setdefault('model_version', model_runmodel_request[0][1])
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })
        # Adding model to the model cache.
        if 'Adding model to the model cache' in log_details and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })
        # Initializing the model process
        if 'Initializing the model process' in log_details and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })

        # Starting model process: \"/home/.venv-default/bin/python -m instabase.model_service.model_runner --registry_import_path /home/ibuser/models/tmpnarhe368 --class_import_path model.model.Codelabs --grpc_server_port 41067 --num_grpc_workers 50 --max_concurrent_rpcs 50\"
        if 'Starting model process: ' in log_details and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })

        # Loading the model
        if 'Loading the model' in log_details and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })

        # Finished model inference in 129.275 secs
        model_inference_time = re.findall(
            r"^Finished model inference in (.*)", log_details
        )
        if model_inference_time and "taskId" in log and 'taskId' in log:
            spans = span_interference_dict.setdefault(log["taskId"], {})
            span = spans.setdefault(log['spanId'], {})
            model_spans = span.setdefault('steps', [])
            model_spans.append({
                'name': log_details,
                'end_time': ts
            })

        if "step" in log:
            # Remove blank step elements that come in from csv to json conversion.
            if log['step'] == '':
                continue

            task_id = log["taskId"]
            tasks_dict = job_details.setdefault("tasks", {})

            step_name = log["step"]
            step_ts = log["ts"]
            step_log = log["log"]

            # Mapping task id to filename
            if task_id in stage_to_filename:
                task_id_dict = tasks_dict.setdefault(
                    stage_to_filename[task_id].split('.')[0], {})
            else:
                task_id_dict = tasks_dict.setdefault(task_id, {})

            steps_dict = task_id_dict.setdefault("steps", {})
            step = steps_dict.setdefault(step_name, {})

            # Calculate time of task
            task_start_time = task_id_dict.setdefault("start_time", ts)
            task_end_time = task_id_dict.setdefault("end_time", ts)

            if task_start_time > ts:
                task_id_dict["start_time"] = ts

            if task_end_time < ts:
                task_id_dict["end_time"] = ts

            task_id_dict["elapsed_time"] = (
                task_id_dict["end_time"] - task_id_dict["start_time"]
            ).total_seconds()

            # Detecting Refiner names in Flowv2
            refiner_name = re.findall(r"^pred:\s(.*)", log_details)
            if refiner_name:
                task_id_dict["refiner_name"] = refiner_name[0]

            # Map Stage to Step ID
            if step_name in stage_to_type:
                step.setdefault("type", stage_to_type[step_name])
                if stage_to_type[step_name] == "refiner" or stage_to_type[step_name] == "apply_refiner":
                    if task_id in span_interference_dict:
                        step['model-service'] = span_interference_dict[task_id]

            elif "-map-" in task_id:
                step.setdefault("type", step_name)
            else:
                step.setdefault("type", "unknown")

            # Calculate model service timestamps.
            if 'model-service' in step:
                for spanid, span in step['model-service'].items():
                    span.setdefault('start_time', span['steps'][0]['end_time'])
                    span.setdefault(
                        'start_time', span['steps'][-1]['end_time'])
                    span.setdefault(
                        'elapsed_time', (span['steps'][-1]['end_time'] - span['steps'][0]['end_time']).total_seconds())
                    for idx, model_steps in enumerate(span['steps']):
                        if idx == 0:
                            continue
                        span['steps'][idx]['start_time'] = span['steps'][idx-1]['end_time']
                        span['steps'][idx]['end_time'] = model_steps['end_time']
                        span['steps'][idx]['elapsed_time'] = (
                            span['steps'][idx]['end_time'] - span['steps'][idx]['start_time']).total_seconds()

            # Step ts
            step_time = datetime.strptime(step_ts, format)
            start_time = step.setdefault("start_time", step_time)
            end_time = step.setdefault("end_time", step_time)

            # Calculate Step start time
            if start_time > step_time:
                step["start_time"] = step_time
            # Calculate Step end time
            if end_time < step_time:
                step["end_time"] = step_time
            # Calculate Step total time taken
            step["elapsed_time"] = (
                step["end_time"] - step["start_time"]
            ).total_seconds()

            refiner = re.findall(
                r"^Running\s(.*)\sfor field\s(.*)\stook\s(.*)", step_log
            )
            if refiner:
                refiner_list = step.setdefault("refiner_fields", [])
                refiner_name = refiner[0][1]
                refiner_time = refiner[0][2]
                refiner_fn = refiner[0][0]
                refiner_dict = {
                    "name": refiner_name,
                    "time_in_ms": refiner_time.replace(" ms", ""),
                    "fn": refiner_fn,
                }

                refiner_list.append(refiner_dict)

            refiner_22_10 = re.findall(
                r"Refiner Record: (.*)\.\w+ Field: (.*) run time: (.*) sec",
                step_log,
            )
            if refiner_22_10:
                refiner_list = step.setdefault("refiner_fields", [])
                refiner_name = refiner_22_10[0][1]
                refiner_time = refiner_22_10[0][2]
                refiner_fn = ""
                refiner_dict = {
                    "name": refiner_name,
                    "time_in_ms": float(refiner_time) * 1000,
                    "fn": refiner_fn,
                }

                refiner_list.append(refiner_dict)

            refiner_2 = re.findall(r"^Field\s(.*)\stook\s(.*)", step_log)
            if refiner_2:
                refiner_list = step.setdefault("refiner_fields", [])
                refiner_name = refiner_2[0][0]
                refiner_time = refiner_2[0][1]
                refiner_fn = ""
                refiner_dict = {
                    "name": refiner_name,
                    "time_in_ms": refiner_time.replace(" ms", ""),
                    "fn": refiner_fn,
                }
                refiner_list.append(refiner_dict)

            # OCR Details
            rotation_time = re.findall(
                r"Selecting best rotation took .*(\d+.\d+s) on (.*)\.\w*_p(\d+)",
                step_log,
            )
            if rotation_time:
                ocr_detail = step.setdefault("pages", {})
                filename = ocr_detail.setdefault(rotation_time[0][1], {})
                pages = filename.setdefault("pages", {})
                page_details = pages.setdefault(rotation_time[0][2], {})
                page_details.setdefault(
                    "rotation_time", rotation_time[0][0])
                step["pages"][rotation_time[0][1]]["pages"] = OrderedDict(
                    sorted(pages.items(), key=lambda t: int(t[0]))
                )


            ocr_time = re.findall(
                r"Request to OCR.*took (\d+.\d+)s on (.*)\.\w*_p(\d+)", step_log
            )
            if ocr_time:
                ocr_detail = step.setdefault("pages", {})
                filename = ocr_detail.setdefault(ocr_time[0][1], {})
                pages = filename.setdefault("pages", {})
                page_details = pages.setdefault(ocr_time[0][2], {})
                page_details.setdefault("ocr_time", ocr_time[0][0])
                step["pages"][ocr_time[0][1]]["pages"] = OrderedDict(
                    sorted(pages.items(), key=lambda t: int(t[0]))
                )
            
            convert_time = re.findall(
                r'(\d+.\d+s) convert_image on (.*)\.\w*_p(\d+)', step_log
            )

            if convert_time:
                ocr_detail = step.setdefault("pages", {})
                filename = ocr_detail.setdefault(convert_time[0][1], {})
                pages = filename.setdefault("pages", {})
                page_details = pages.setdefault(convert_time[0][2], {})
                page_details.setdefault("convert_time", convert_time[0][0])
                step["pages"][convert_time[0][1]]["pages"] = OrderedDict(
                    sorted(pages.items(), key=lambda t: int(t[0]))
                )

            # writing_pdf_to_ibocr_time = re.findall(
            #     r"(^.*) Writing\s.*ibocr.*Path: (.*)", step_log
            # )
            # if writing_pdf_to_ibocr_time:
            #     ocr_detail = step.setdefault('pages', {})
            #     detail = ocr_detail.setdefault('pdf_to_ibocr_conv', {})
            #     detail.setdefault('writing_time', writing_pdf_to_ibocr_time[0][0] )

            # writing_ibocr_to_csv_time = re.findall(
            #     r"(^.*) writing ibocr as csv", step_log
            # )
            # if writing_ibocr_to_csv_time:
            #     ocr_detail = step.setdefault('pages', {})
            #     detail = ocr_detail.setdefault('ibocr_to_csv', {})
            #     detail.setdefault('writing_time', writing_ibocr_to_csv_time[0] )

            # Update Reduce UDF Sections as filenames stretch a single stage
            if step_name in reduce_stage_to_filename:
                temp_reduce_udf = reduce_stage_to_filename[step_name]
                for reduce_file_name in temp_reduce_udf:
                    if reduce_file_name in job_details["tasks"]:
                        temp_step = job_details["tasks"][reduce_file_name][
                            "steps"
                        ].setdefault(step_name, {})
                        temp_step.update(step)
    
    #Calculate out waiting time for app-task
    for taskName, taskValues in job_details["tasks"].items():
        previous_time = []
        for stepName, stepValues in taskValues["steps"].items():
            print(taskValues)
            if previous_time:
                step_start_time = stepValues['start_time']
                step_end_time = stepValues['end_time']

                #How many steps to go back to find link for waiting for resources
                timeback_steps = 3
                for timeback in range(1,timeback_steps):
                    if previous_time[-abs(timeback)] > step_start_time:
                        continue
                    else:
                        true_previous_time = previous_time[-abs(timeback)]
                        break
                waiting_time = (step_start_time - true_previous_time).total_seconds()
                if waiting_time >= waiting_threshold:
                    job_details['tasks'][taskName]['steps'][stepName]['waiting_for_resources'] = {
                        'start_time': true_previous_time,
                        'end_time' : step_start_time,
                        'elapsed_time' : waiting_time
                    }
                previous_time.append(step_end_time)
            else:
                previous_time.append(stepValues['end_time'])
    
    # Clean up any Datatime obj in dict
    return json.loads(json.dumps(job_details, indent=2, default=str))







def aggergate_details(job_details):
    job_details = job_details.copy()
    filenames = job_details.setdefault("filenames", [])
    ocr_time = job_details.setdefault("ocr_time", [])
    ocr_detail_2 = job_details.setdefault("ocr_detail", [])
    refiner_details = job_details.setdefault("refiner_details", [])
    files_dict = job_details.setdefault("files", [])
    steps = job_details.setdefault("steps", [])
    model_service_details = job_details.setdefault('model_service', [])

    for k, v in job_details["tasks"].items():
        # files_dict.setdefault(v['filename'], v['elapsed_time'])
        if "Stage" not in k:
            filenames.append(k)
        details = []
        for x, y in v["steps"].items():
            step_dict = {
                "job_id": job_details['jobid'],
                "filename": k,
                "step_id": x,
                "type": y["type"],
                "start_time": y["start_time"],
                "end_time": y["end_time"],
                "elapsed_time": y["elapsed_time"],
            }
            steps.append(step_dict)
            sub_details = []
            type = y["type"]
            if "refiner_fields" in y:
                for field in y["refiner_fields"]:
                    field_details = {
                        "job_id": job_details['jobid'],
                        "filename": k,
                        "step_id": x,
                        "time_in_s": float(field["time_in_ms"]) / 1000,
                    }
                    field_details.update(field)
                    refiner_details.append(field_details)
                sub_details = y["refiner_fields"]

            if 'model-service' in y:
                for spans_idx, spans in y['model-service'].items():
                    for span_step in spans['steps']:
                        model_service_details.append({
                            "job_id": job_details['jobid'],
                            "filename": k,
                            "step_id": x,
                            "name": spans['model_name'] if 'model_name' in spans else None,
                            "version": spans['model_version'] if 'model_version' in spans else None,
                            "action": span_step['name'],
                            "start_time": span_step['start_time'] if 'start_time' in span_step else None,
                            "end_time": span_step['end_time'],
                            "elapsed_time": span_step['elapsed_time'] if 'elapsed_time' in span_step else None
                        })
# "end_time": "2023-09-11 05:02:32",["Even" if i%2==0 else "Odd" for i in range(10)]
#                   "start_time": "2023-09-11 05:02:32",
#                   "elapsed_time": 0.0
            if "pages" in y:
                new_pages = []
                for file_name, pages_details in y["pages"].items():
                    for page_num, page_details in pages_details["pages"].items():
                        ocr_time.append(float(page_details["ocr_time"]))

                        if 'convert_time' in page_details:
                            convert_time = float(page_details['convert_time'].replace("s", ""))
                        else:
                            convert_time = ''
                        if 'rotation_time' in page_details:
                            rotation_time = float(page_details['rotation_time'].replace("s", ""))
                        else:
                            rotation_time = ''

                        ocr_detail_2.append(
                            {"job_id": job_details['jobid'],
                                "filename": k,
                                "step_id": x,
                                "page_num": page_num,
                                "ocr_time": float(page_details["ocr_time"]),
                                "convert": convert_time,
                                "rotation": rotation_time,
                            }
                        )

                        new_pages.append(
                            {
                                "page_num": "pg " + page_num,
                                "ocr_time": page_details["ocr_time"],
                            }
                        )
                    break
                sub_details = new_pages

            if 'waiting_for_resources' in y:
                waiting_for_resources = y['waiting_for_resources']
                details.append({
                    "x" : [waiting_for_resources['start_time'], waiting_for_resources['end_time']],
                    'y' : f'{x}_waiting',
                    'type': 'waiting',
                    'elapsed_time' : waiting_for_resources['elapsed_time']
                    })
            details.append(
                {
                    "x": [y["start_time"], y["end_time"]],
                    "y": x,
                    "elapsed_time": y["elapsed_time"],
                    "type": type,
                    "sub_details": sub_details,
                }
            )
        if 'Stage' in k and y["elapsed_time"] == 0:
            continue
        files_dict.append(
            {
                "x": [v["start_time"], v["end_time"]],
                "y": k,
                "elapsed_time": y["elapsed_time"],
                "type" : 'file',
                "detail": details,
            }
        )

    job_details.setdefault(
        "ocr_average", round(numpy.average(ocr_time), 2))
    job_details.setdefault("ocr_count", len(ocr_time))
    job_details.setdefault(
        "longest_step", max(steps, key=itemgetter("elapsed_time"))
    )
    job_details.setdefault('job_details', {
        'job_id': job_details['jobid'],

        'ocr_average': job_details['ocr_average'],
        'ocr_count': job_details['ocr_count'],
        'start_time': job_details['start_time'],
        'end_time': job_details['end_time'],
        'elapsed_time': job_details['elapsed_time']
    })
    return job_details


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help='Input can be either a single file or a folder')
    parser.add_argument("-o", "--outputDir", required=True, help='Output location will be created automatically based on value')
    parser.add_argument("--output_type", nargs='+', default = 'json', help="excel | json | html")
    parser.add_argument("--waiting-threshold", default=1, help='Threshold to indicate waiting for resources in seconds')
    args = parser.parse_args()

    waiting_threshold = float(args.waiting_threshold)
    aggergate_job_details = []
    outputDir = Path(args.outputDir)
    if not outputDir.exists():
        os.mkdir(outputDir)

    files_to_parse = []
    if os.path.isfile(args.input):
        files_to_parse.append(args.input)
    else:
        for filename in os.listdir(args.input):
            t = os.path.join(args.input, filename)
            if os.path.isfile(t) :
                files_to_parse.append(t)

    if 'excel' in args.output_type:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        combined_jobs_df = None
        combined_steps_df = None
        combined_refiner_details_df = None
        combined_ocr_details_df = None

    for file in files_to_parse:
        if file.endswith("json"):
            log_detail = json.load(open(file))
        elif file.endswith("csv"):
            reader = csv.DictReader(open(file))
            log_detail = list()
            for items in reader:
                log_detail.append(items)
        else:
            continue
        

        job_details = parse(log_detail, waiting_threshold)
        aggergate_job_details.append(aggergate_details(job_details))

        if 'html' in args.output_type:
            job_details_agg = aggergate_details(job_details)
            p = Path(__file__).parent.parent
            template_path = f'{p}/flowparser/templates/'
            environment = Environment(loader=FileSystemLoader(template_path))
            template = environment.get_template("template.html")
            html_file = template.render(job=job_details_agg)
            html_output = open(
                f"{outputDir}/{job_details['jobid']}.html", "w")
            html_output.write(html_file)

        if 'json' in args.output_type:
            # job_details_agg = aggergate_details(job_details)
            output = open(f"{outputDir}/{job_details['jobid']}.json", "w")
            output.write(json.dumps(
                job_details, indent=2, default=str))
            output.close()

        if 'excel' in args.output_type:

            job_details_agg = aggergate_details(job_details)
            jobs_df = pd.DataFrame.from_dict([
                job_details_agg['job_details']])
            steps_df = pd.DataFrame.from_dict(job_details_agg['steps'])
            ocr_details_df = pd.DataFrame.from_dict(
                job_details_agg['ocr_detail'])
            refiner_details_df = pd.DataFrame.from_dict(
                job_details_agg['refiner_details'])

            if combined_jobs_df is None:
                combined_jobs_df = jobs_df.copy()
            else:
                combined_jobs_df = pd.concat(
                    [combined_jobs_df, jobs_df], axis=0)

            if combined_steps_df is None:
                combined_steps_df = steps_df.copy()
            else:
                combined_steps_df = pd.concat(
                    [combined_steps_df, steps_df], axis=0)

            if combined_refiner_details_df is None:
                combined_refiner_details_df = refiner_details_df.copy()
            else:
                combined_refiner_details_df = pd.concat(
                    [combined_refiner_details_df, refiner_details_df], axis=0)

            if combined_ocr_details_df is None:
                combined_ocr_details_df = ocr_details_df.copy()
            else:
                combined_ocr_details_df = pd.concat(
                    [combined_ocr_details_df, ocr_details_df], axis=0)

            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            jobs_sheet = wb.create_sheet('Jobs')
            steps_sheet = wb.create_sheet('Steps')
            ocr_details_sheet = wb.create_sheet('OCR Details')
            refiner_details_sheet = wb.create_sheet('Refiner Details')

            for r in dataframe_to_rows(jobs_df, index=True, header=True):
                jobs_sheet.append(r)
            for r in dataframe_to_rows(steps_df, index=True, header=True):
                steps_sheet.append(r)
            for r in dataframe_to_rows(ocr_details_df, index=True, header=True):
                ocr_details_sheet.append(r)
            for r in dataframe_to_rows(refiner_details_df, index=True, header=True):
                refiner_details_sheet.append(r)

            wb.save(f"{outputDir}/{job_details_agg['jobid']}.xlsx")

    # Aggregation of data
    if 'json' in args.output_type:
        output = open(f"{outputDir}/job_aggregation.json", "w")
        output.write(json.dumps(aggergate_job_details, indent=2, default=str))
        output.close()

    if 'excel' in args.output_type:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        jobs_sheet = wb.create_sheet('Jobs')
        steps_sheet = wb.create_sheet('Steps')
        ocr_details_sheet = wb.create_sheet('OCR Details')
        refiner_details_sheet = wb.create_sheet('Refiner Details')

        for r in dataframe_to_rows(combined_jobs_df, index=True, header=True):
            jobs_sheet.append(r)
        for r in dataframe_to_rows(combined_steps_df, index=True, header=True):
            steps_sheet.append(r)
        for r in dataframe_to_rows(combined_ocr_details_df, index=True, header=True):
            ocr_details_sheet.append(r)
        for r in dataframe_to_rows(combined_refiner_details_df, index=True, header=True):
            refiner_details_sheet.append(r)

        wb.save(f"{outputDir}/job_aggregation.xlsx")

if __name__ == "__main__":
    main()
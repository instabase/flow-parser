"""
MIT License

Copyright (c) 2023 Instabase

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
import json
from ..exceptions import *
import os
from pathlib import Path
from jinja2 import Environment, FileSystemLoader
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class Render(object):
    def __init__(self, job, output_dir, multi_job=False):
        super(Render, self).__init__()
        self.job = job
        self.multi_job = multi_job
        self.output_dir = output_dir
    
    def file_writer(self, name, data):
        try:
            if not self.output_dir.exists():
                os.mkdir(self.output_dir)

            file_obj = open(f'{self.output_dir}/{name}', 'w')
            file_obj.write(data)
            file_obj.close()
        except FileNotFoundError as e:
            raise FileWriterError(name)

    def format_job(self, job, default=str):
        return json.dumps(job, indent=2, default=default)

    def json(self):
        job = self.job
        if self.multi_job:
            job_id = 'multi_job'
        else:
            job_id = job['jobid']
        self.file_writer(f'{job_id}.json', self.format_job(job))

    def html(self):
        job = self.job
        job_id = job['jobid']
        p = Path(__file__).parent.parent
        
        template_path = f'{p}/utils/templates'
        environment = Environment(loader=FileSystemLoader(template_path))
        template = environment.get_template("template.html")
        html_file = template.render(job=job)
        self.file_writer(f'{job_id}.html', html_file)

    def excel(self):
        job = self.job
        job_id = job['jobid']
        jobs_df = pd.DataFrame.from_dict([
            job['job_details']])
        steps_df = pd.DataFrame.from_dict(job['steps'])
        ocr_details_df = pd.DataFrame.from_dict(
            job['ocr_detail'])
        refiner_details_df = pd.DataFrame.from_dict(
            job['refiner_details'])

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        jobs_sheet = wb.create_sheet('Jobs')
        steps_sheet = wb.create_sheet('Steps')
        ocr_details_sheet = wb.create_sheet('OCR Details')
        refiner_details_sheet = wb.create_sheet('Refiner Details')

        for r in dataframe_to_rows(jobs_df, index=True, header=True):
            jobs_sheet.append(r)
        for r in dataframe_to_rows(steps_df, index=True, header=True):
            steps_sheet.append(r)
        for r in dataframe_to_rows(ocr_details_df, index=True, header=True):
            ocr_details_sheet.append(r)
        for r in dataframe_to_rows(refiner_details_df, index=True, header=True):
            refiner_details_sheet.append(r)

        wb.save(f"{self.output_dir}/{job_id}.xlsx")


    
